from datetime import datetime
from pytz import timezone
import os
from odoo import api, fields, models, _
from odoo.exceptions import UserError


class HHExpenseInputAnticipatedUploadDate(models.Model):
    _name = "hhexpense.anticipated.date"
    _rec_name = 'name'

    name = fields.Date(string='Anticipated HSBC Upload Date')
    has_input = fields.Boolean(default=False)
    hhexpense_line_id = fields.One2many('hhexpense.line', inverse_name='anticipated_date_id')
    create_input_date = fields.Date(readonly=True, default=fields.datetime.now())
    upload_attempt_counter = fields.Integer(compute="_compute_generate_upload_attempt_counter", store=True)
    # modification_counter = fields.Char(default='00')

    @api.onchange('name')
    def _onchange_name(self):
        # Purpose of this function:
        # 1.Check if user input is appropriate
        # 2.Show warning message based on different wrong input situation

        # User has input
        if self.name:
            # Get user input value, the data type is <class 'str'>
            input_date = self.name
            # Convert input to data type --- date
            # if not using ".date()", the data type is <class 'datetime.datetime'>,
            # we need <class 'datetime.date'> for comparison
            convert_input = datetime.strptime(input_date, "%Y-%m-%d").date()
            print("User's input(already converted to proper type) --- Anticipated Upload Date is:: ", convert_input)
            # Today is?
            # if doing this: today = datetime.now(),
            # the data type will be: type:<class 'datetime.datetime'>, value is like: 2018-08-08 09:22:21.557078,
            # we need <class 'datetime.date'> for comparison
            today = datetime.now().date()
            # Show different message if input date smaller than today OR it is not a valid date
            if convert_input < today:
                print("Input smaller than today")
                return {
                    'warning': {
                        'title': "Incorrect input",
                        'message': "Anticipated upload date can't be later than today("
                                   + str(today)
                                   + "), Please select again",
                    }
                }
            else:
                # Obtain all not valid date (from other model) and save it into list for comparison
                not_valid_date = self.env['hhexpense.holiday.date'].search([])
                not_valid_date_list = []
                for date in not_valid_date:
                    # print("get data: ", type(date.holiday_date))
                    not_valid_date_list.append(date.holiday_date)
                # print("Show all non-valid date :", not_valid_date_list)
                # Before accept input value, check if this date is business day (not holiday date and weekends)?
                if input_date in not_valid_date_list:
                    self.has_input = False
                    print("Input is not a business day")
                    return {
                        'warning': {
                            'title': "Incorrect date",
                            'message': "Anticipated upload date should be a business day, Please select again",
                        }
                    }
                else:
                    self.has_input = True
                    print("Input is appropriate, Now you can see button and call auto_processing function!")
                    # self.function_one()
        # No user input
        else:
            # print("user didnt input yet")
            pass

    @api.depends('name')
    def _compute_generate_upload_attempt_counter(self):
        # this function will be used to generate an number to
        # store "attempting download times of the HSBC file" for the same day.
        # number's range is 0-9, which means accountant has 10 chances to download HSBC file at same day

        print("Generating attempting number")

        # Use list to store today's "attempting download times" information
        today_attempt_times_list = []
        today_date = int(datetime.today().strftime('%Y%m%d'))
        # Search existing record from database
        anticipated_date_model_data = self.env['hhexpense.anticipated.date'].search([])
        for rec in anticipated_date_model_data:

            # if rec.name:
            #     model_has_record = True
            # else:
            #     model_has_record = False
            print("date that user download file (int) :", int(rec.create_input_date.replace('-', '')))
            print("today's date (int)                 :", today_date)

            # We only need today's "attempting download times" value
            # get those value by using "create_input_date" as a filter to compare with today's date - "current_date".
            # Thus, for comparison, we need modify "create_input_date" value into correct format(2018-09-31 to 20180931)
            if int(rec.create_input_date.replace('-', '')) == today_date:
                today_attempt_times_list.append(rec.upload_attempt_counter)
            else:
                print("you haven't try to download file today")

        print("today's attempting times list          :", today_attempt_times_list)

        if len(today_attempt_times_list) > 1:
            self.upload_attempt_counter = max(today_attempt_times_list) + 1
            print("not today's first time download :", self.upload_attempt_counter)
        else:
            self.update({'upload_attempt_counter': 0})
            print("today's first time download")

        # make sure user will not download more than 9 times
        if self.upload_attempt_counter > 9:
            raise UserError(_("You already attempted to download file more than 9 times a day!"))
        else:
            print("You still have chance to download")


    def detect_checked_records(self):
        # Find out selected record by reading id file and we will save into a list

        selected_rec_id_file = open("./select_rec.txt", "r", encoding="utf-8")

        selected_rec_id_list = []
        for line in selected_rec_id_file:
            pure_data = line.strip('\n')
            selected_rec_id_list.append(pure_data)
        print("selected record id list: ", selected_rec_id_list)
        selected_rec_id_file.close()
        os.remove("./select_rec.txt")
        return selected_rec_id_list

    # def auto_processing(self):
    #     # Purpose of this function:
    #     # 1.Update anticipated date to selected record and Mark record as processing
    #     # 2.Save input date into file, used in controller
    #     # 3.Check if any record ready to post
    #     # 4.Generate HSBC txt file
    #
    #     # 1.
    #     # Find out selected record by reading id file and we will save into a list
    #     # selected_rec_id_file = open("./hsbc_intermediate_file_get_selected_rec_id.txt", "r", encoding="utf-8")
    #     # selected_rec_id_list = []
    #     # for line in selected_rec_id_file:
    #     #     pure_data = line.strip('\n')
    #     #     selected_rec_id_list.append(pure_data)
    #     # print("selected record id list: ", selected_rec_id_list)
    #     self.detect_checked_records()
    #     selected_rec_id_list = self.detect_checked_records()
    #     # Get all record from hhexpense_line model and update user input (anticipated date) to selected record only
    #     all_sub_rec = self.env['hhexpense.line'].search([])
    #     for rec in all_sub_rec:
    #         # "rec.id" data type is int, thus, we need convert to string for comparison
    #         rec_id = str(rec.id)
    #         # Updating anticipated date to selected record and mark it as processing
    #         if rec_id in selected_rec_id_list:
    #             rec.update({
    #                 'anticipated_date': self.name,
    #                 'expense_line_is_processing': True
    #             })
    #             print("Updating this record！", self.name)
    #         else:
    #             print("Ignore this record!")
    #             pass
    #
    #     # 2.
    #     # Selected record's anticipated_date file
    #     temp_file_get_anticipated_date = open("./hsbc_intermediate_file_get_selected_rec_anticipated_date.txt", "w",
    #                                           encoding="utf-8")
    #     temp_file_get_anticipated_date.write(self.name + '\n')
    #     temp_file_get_anticipated_date.close()
    #
    #     # 3.
    #     # self.check_if_expenses_allowed_to_post()
    #
    #     # 4.
    #     return {
    #         'name': 'Go to website',
    #         'type': 'ir.actions.act_url',
    #         'target': 'self',
    #         'url': '/web/binary/generate_hsbcnet_txt_file'
    #     }
    #
    # # def check_if_expenses_allowed_to_post(self):
    # #     # Purpose of this function:
    # #     # If allowed, record will be moved to paid and ready for posting
    # #     # i.e. Under the same master record, this expense can be post only when
    # #     # all sub bank records is processing and all sub cash records is paid
    # #     print("trigger by bank action")
    # #     # get all sub rec
    # #     master_records = self.env['hhexpense.hhexpense'].search([])
    # #     # Go through each master record in database
    # #     for rec in master_records:
    # #         print("master rec:   ", rec.name)
    # #         print("count sub rec:", len(rec.expense_line))
    # #         # For each sub item in this master record
    # #         for sub_rec in rec.expense_line:
    # #             if sub_rec.state == "verified":
    # #                 if sub_rec.pay_bank_cash == "Bank":
    # #                     if not sub_rec.expense_line_is_processing:
    # #                         print("bank not match")
    # #                         break
    # #                 elif sub_rec.pay_bank_cash == "Cash":
    # #                     if not sub_rec.expense_line_is_paid:
    # #                         print("cash not match")
    # #                         break
    # #         # no break during for-loop, thus, this record can be output. i.e. mark it as paid, ready to post
    # #         else:
    # #             rec.state = "paid"
    # #             print("this expense is ready to post")
    # #     print("trigger by bank action --- done")

    def generate_reports(self, **kwargs):
        import shutil
        from openpyxl import load_workbook
        import os

        if not os.path.exists("./select_rec.txt"):  # ensure the download reports button is pressed once
            # No pop up message appear now. Error log not seen too
            raise UserError("You have already downloaded the reports.")

        # Generate a batch number first
        batch_no = self.generate_batch_number()
        has_file = ''
        file_list = [['./expense_excel_report_rmb.xlsx', './hsbc_rmb.txt', 'R'],
                     ['./expense_excel_report_hkd.xlsx', './hsbc_hkd.txt', 'H']]

        def edit_excel(file, cur):
            # add batch number to exl report and rename file
            wb = load_workbook(file)
            ws = wb.worksheets[0]
            ws['A5'] = "BATCH NO.: " + batch_no
            wb.save(file)
            shutil.move(file, './%s%s.xlsx' % (cur, batch_no))

        def edit_text(file, cur):
            txt = open(file, 'r', newline='\r\n')
            content = txt.read()
            txt.close()
            # add anticipated date to txt file
            # overwrite the existing file and rename to avoid creating multiple files
            txt = open(file, 'w', encoding='utf-8')
            date = self.name  # the anticipated upload date
            anti_date = date[8:] + date[5:7] + date[2:4]
            content = content.replace('AnTiDa', str(anti_date))
            att_no = 'Ex' + str(self.upload_attempt_counter)
            content = content.replace('Exp', att_no, 1)
            txt.write(content)
            txt.close()
            shutil.move(file, './%s%s.txt' % (cur, batch_no))

        for files in file_list:
            if os.path.exists(files[0]):
                cur = ((files[0].split('_')[-1])[:3]).upper() + '_'  # from rmb to RMB_
                edit_excel(files[0], cur)
                edit_text(files[1], cur)
                has_file = has_file + files[2]

        return {
            'type': 'ir.actions.act_url',
            'target': 'self',
            'url': '/web/binary/download_reports?batch_no=%s&has_file=%s' % (batch_no, has_file)
        }

    def generate_batch_number(self):
        # 2-- generate batch_number without bank/cash type and modification_counter
        # the format of batch_number is [year-month-day-hour + A + modification_counter]
        # today = datetime.now()
        raw_batch_number = datetime.now(tz=timezone('Asia/Hong_Kong')).strftime('%Y%m%d%H%M')
        print('batch_number', raw_batch_number)

        # 3-- get all selected records from selected_rec_id_list
        expense_list = []
        selected_rec_id_list = self.detect_checked_records()
        all_sub_rec = self.env['hhexpense.line'].search([])
        # 3.1-- check info for selected records
        for rec in all_sub_rec:
            # "rec.id" data type is int, thus, we need convert to string for comparison
            rec_id = str(rec.id)
            # 3.2-- check whether record already has modification_number
            if rec_id in selected_rec_id_list:
                rec.update({
                    'has_batch_number': True,
                    'batch_number': raw_batch_number
                })
                # also update its expense level's data "batch_number_copy"
                rec.expense_id.batch_number_copy = raw_batch_number
                print("Generated a batch number for this record      :", rec.batch_number)
                print("Generated a batch number for its master record:", rec.expense_id.batch_number_copy)
                if rec.expense_id.id not in expense_list:
                    expense_list.append(rec.expense_id.id)

            else:
                print("Ignore this record!")
                pass
        print(expense_list)
        self.env['hhexpense.hhexpense'].search([('id', '=', expense_list[-1])]).review_expense_email()
        return raw_batch_number


class HHExpenseInputPaymentApprovedDate(models.Model):
    _name = "hhexpense.payment.date"

    name = fields.Date(string='Payment Date')
    has_input = fields.Boolean(default=False)
    hhexpense_line_id = fields.One2many('hhexpense.line', inverse_name='payment_approved_date_id', )

    # @api.multi
    # def generate_flexacc_txt_file(self):
    #     print("generate_flexacc_txt_file function called")

    @api.onchange('name')
    def _onchange_name(self):
        # Purpose of this function:
        # 1.Check if user input is appropriate
        # 2.Show warning message

        # User has input
        if self.name:
            # Get user input value, the data type is <class 'str'>
            input_date = self.name
            # Convert input to data type --- date
            # if not using ".date()", the data type is <class 'datetime.datetime'>,
            # we need <class 'datetime.date'> for comparison
            convert_input = datetime.strptime(input_date, "%Y-%m-%d").date()
            print("User's input(already converted to proper type) --- Payment Approve Date is:: ", convert_input)
            # Today is?
            # if doing this: today = datetime.now(),
            # the data type will be: type:<class 'datetime.datetime'>, value is like: 2018-08-08 09:22:21.557078,
            # we need <class 'datetime.date'> for comparison
            today = datetime.now().date()
            # Show different message if input date smaller than today OR it is not a valid date
            if convert_input < today:
                self.has_input = False
                print("Input smaller than today")
                return {
                    'warning': {
                        'title': "Incorrect input",
                        'message': "payment approve date can't be later than today("
                                   + str(today)
                                   + "), Please select again",
                    }
                }
            #
            else:
                self.has_input = True
        # No user input
        else:
            # print("user didnt input yet")
            pass

    def auto_post_processing(self):
        # Purpose of this function:
        # 1.Update payment approve date to selected record and Mark record as posted
        # send out email
        # 2.Save input date into file, used in controller
        # 3.Ask controller to generate FlexAcc txt file, i.e. return url to controller

        # 1.

        # "do not send second time" control
        # use a list to store all expense_id (master record),
        # so that the times of sending email will be equal to number of master records
        master_record_id_set = set()

        # Find out selected record by reading id file and we will save into a list
        selected_rec_id_file = open("./flexacc_intermediate_file_get_selected_rec_id.txt", "r", encoding="utf-8")
        selected_rec_id_list = []
        for line in selected_rec_id_file:
            pure_data = line.strip('\n')
            selected_rec_id_list.append(pure_data)
        print("selected record id list: ", selected_rec_id_list)
        # Get all record from hhexpense_line model
        # and update user input (payment approve date) to selected record only
        all_sub_rec = self.env['hhexpense.line'].search([])
        for rec in all_sub_rec:
            # "rec.id" data type is int, thus, we need convert to string for comparison
            rec_id = str(rec.id)
            # Updating payment approve date to selected record and mark it as posted
            if rec_id in selected_rec_id_list:
                rec.expense_id.state = "posted"
                rec.update({
                    'payment_approved_date': self.name,
                    # 'state': 'posted',
                })
                print(rec.expense_id.payment_approved_date)
                if rec.expense_id.payment_approved_date is False:
                    rec.expense_id.update({
                        'payment_approved_date': self.name,
                    })
                else:
                    pass
                print("Updating this record！", self.name)

                # add non-repeat master record's id into set
                master_record_id_set.add(rec.expense_id)

            else:
                print("Ignore this record!")
                pass

        # send email for those master record only (email base --- master record)
        for mst_rec_id in master_record_id_set:
            print("mst_rec_id: ", mst_rec_id)
            mst_rec_id.pay_expense()
            print("send")

        # 2.
        # Selected record's payment_approve_date file
        temp_file_get_payment_approve_date = open(
            "./flexacc_intermediate_file_get_selected_rec_payment_approve_date.txt",
            "w", encoding="utf-8"
        )
        temp_file_get_payment_approve_date.write(self.name + '\n')
        temp_file_get_payment_approve_date.close()

        # 3.
        return {
            'name': 'Go to website',
            'type': 'ir.actions.act_url',
            'target': 'self',
            'url': '/web/binary/generate_flexacc_txt_file'
        }

    # Following code is for old workflow (expense level's post action)
    # def auto_post_processing(self):
    #     # Purpose of this function:
    #     # 1.Update payment approve date to selected record and Mark record as posted
    #     # 2.Save input date into file, used in controller
    #     # 3.Ask controller to generate FlexAcc txt file, i.e. return url to controller
    #
    #     # 1.
    #     # Find out selected record by reading id file and we will save into a list
    #     selected_rec_id_file = open("./flexacc_intermediate_file_get_selected_rec_id.txt", "r", encoding="utf-8")
    #     selected_rec_id_list = []
    #     for line in selected_rec_id_file:
    #         pure_data = line.strip('\n')
    #         selected_rec_id_list.append(pure_data)
    #     print("selected record id list: ", selected_rec_id_list)
    #     # Get all record from hhexpense_hhexpense model
    #     # and update user input (payment approve date) to selected record only
    #     all_master_rec = self.env['hhexpense.hhexpense'].search([])
    #     for rec in all_master_rec:
    #         # "rec.id" data type is int, thus, we need convert to string for comparison
    #         rec_id = str(rec.id)
    #         # Updating payment approve date to selected record and mark it as posted
    #         if rec_id in selected_rec_id_list:
    #             rec.update({
    #                 'payment_approved_date': self.name,
    #                 'state': 'posted'
    #             })
    #             print("Updating this record！", self.name)
    #         else:
    #             print("Ignore this record!")
    #             pass
    #
    #     # 2.
    #     # Selected record's payment_approve_date file
    #     temp_file_get_payment_approve_date = open(
    #         "./flexacc_intermediate_file_get_selected_rec_payment_approve_date.txt",
    #         "w", encoding="utf-8"
    #     )
    #     temp_file_get_payment_approve_date.write(self.name + '\n')
    #     temp_file_get_payment_approve_date.close()
    #
    #     # 3.
    #     return {
    #         'name': 'Go to website',
    #         'type': 'ir.actions.act_url',
    #         'target': 'self',
    #         'url': '/web/binary/generate_flexacc_txt_file'
    #     }
